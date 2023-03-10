from openpyxl.worksheet.worksheet import Worksheet

MAX_COLUMNS = 41
MIN_COLUMNS = 6
MAX_ROWS = 6
MIN_ROWS = 4

class InvalidTable(Exception):
  "Raised when size in table are not correct"

  def __init__(self, message='El tamaño de la tabla es incorrecto.'):
    self.message = message
    super().__init__(self.message)


def optimum(sheet : Worksheet):
  rows, columns = sheet.max_row, sheet.max_column
  last_process = sheet.cell(column=columns, row = 1).value
  frames, fails, lasted = rows - 2, 0, 0
  previous_row = [None]

  if columns < MIN_COLUMNS: raise InvalidTable(f'El tamaño de la tabla es muy pequeña')
  if columns > MAX_COLUMNS: raise InvalidTable(f'El tamaño máximo de la tabla es {MAX_COLUMNS} columnas')
  if rows < MIN_ROWS: raise InvalidTable(f'Tamaño minimo de marcos son {MIN_ROWS}.')
  if rows > MAX_ROWS: raise InvalidTable(f'Tamaño máximo de marcos es {MAX_ROWS}.')

  for cols in sheet.iter_cols(min_col = 2, max_col = columns, max_row = rows):
    row = []
    for cell in cols:
      if cell.row == 1:
        if (new_value := cell.value) == None:
          raise InvalidTable('La sintaxis de la tabla en el archivo no es válida.')

      if 1 < cell.row < rows:
        row.append(cell.value)

      if cell.row == rows:
        if cell.value is not None:
          fails += 1

    if new_value not in previous_row:
      try:
        if None not in previous_row:
          if previous_row.index(last_process) != row.index(new_value): 
            assert False, 'Este no es el algoritmo optimo'
      except ValueError:
        assert previous_row[lasted % frames] != row[lasted % frames], 'Este no es el algoritmo optimo'
        lasted += 1
    previous_row = row
  
  if fails == 0:
    raise InvalidTable('La sintaxis de la tabla en el archivo no es válida')

  return 'Optimo', columns - 1, fails
      

def fifo(sheet : Worksheet):
  rows, columns = sheet.max_row, sheet.max_column
  frames, fails, lasted = rows - 2, 0, 0
  previous_row = [None]

  for cols in sheet.iter_cols(min_col = 2, max_col = columns, max_row = rows):
    row = []
    for cell in cols:
      if cell.row == 1:
        new_value = cell.value

      if 1 < cell.row < rows:
        row.append(cell.value)

      if cell.row == rows:
        if cell.value is not None:
          fails += 1

    if new_value not in previous_row:
      if None not in previous_row:
        assert new_value == row[lasted % frames], 'Este no es el algoritmo fifo'
        lasted += 1

    previous_row = row

  return 'FIFO', columns - 1, fails

def not_recently_used(sheet : Worksheet):
  rows, columns = sheet.max_row, sheet.max_column
  fails = 0
  previous_row = [None]
  last_used = []

  for cols in sheet.iter_cols(min_col = 2, max_col = columns, max_row = rows):
    row = []
    for cell in cols:
      if cell.row == 1:
        new_value = cell.value
        if new_value in last_used:
          last_used.pop(last_used.index(new_value))
        last_used.append(new_value)

      if 1 < cell.row < rows:
        row.append(cell.value)

      if cell.row == rows:
        if cell.value is not None:
          fails += 1

    if new_value not in previous_row:
      if None not in previous_row:
        deleted = last_used.pop(0)
        assert deleted not in row, 'No es el algoritmo no usadas recientemente'
    previous_row = row

  return 'No usadas recientemente', columns - 1, fails
  

def second_chance(sheet : Worksheet):
  rows, columns = sheet.max_row, sheet.max_column
  frames, fails, lasted = rows - 2, 0, 0
  previous_row, bits = [None], [0] * frames

  for cols in sheet.iter_cols(min_col = 2, max_col = columns, max_row = rows):
    row = []
    for cell in cols:
      if cell.row == 1:
        new_value = cell.value

      if 1 < cell.row < rows:
        row.append(cell.value)

      if cell.row == rows:
        if cell.value is not None:
          fails += 1

    if new_value not in previous_row:
      if None not in previous_row:
        i = lasted
        while True:
          if bits[i % frames] == 0:
            break
          else:
            bits[i % frames] = 0
          i += 1

        if new_value == row[lasted % frames]:
          lasted += 1
        else:
          assert row[i % frames] == new_value, 'Este no es segunda oportunidad'
    else:
      bits[previous_row.index(new_value)] = 1
    
    previous_row = row

  return 'Segunda oportunidad', columns - 1, fails

def clock(sheet : Worksheet):
  rows, columns = sheet.max_row, sheet.max_column
  frames, fails, lasted = rows - 2, 0, 0
  previous_row, bits = [None], [0] * frames

  for cols in sheet.iter_cols(min_col = 2, max_col = columns, max_row = rows):
    row = []
    for cell in cols:
      if cell.row == 1:
        new_value = cell.value

      if 1 < cell.row < rows:
        row.append(cell.value)

      if cell.row == rows:
        if cell.value is not None:
          fails += 1

    if new_value not in previous_row:
      if None not in previous_row:
        i = lasted
        while True:
          if bits[i % frames] == 0:
            break
          else:
            bits[i % frames] = 0
          i += 1

        lasted = i + 1
        assert row[i % frames] == new_value, 'Este no es reloj mejorado'

    else:
      bits[previous_row.index(new_value)] = 1
    previous_row = row

  return 'Reloj mejorado', columns - 1, fails
